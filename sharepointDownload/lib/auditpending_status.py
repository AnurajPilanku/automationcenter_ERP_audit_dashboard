
'''

Author      :      Anuraj Pilanku

Use Case    :     ERP Audit Verification Automation

'''
import openpyxl
import os
import json
import datetime
import shutil
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from os.path import basename
import pandas as pd
import sys
from datetime import date
import datetime
import time
from exchangelib.protocol import BaseProtocol
from exchangelib import Credentials, Account, DELEGATE, Configuration, FileAttachment
import requests
from bs4 import BeautifulSoup, BeautifulStoneSoup

today=str(datetime.datetime.today().date())
passwordUpdate='''"C:\Program Files (x86)\CyberArk\ApplicationPasswordSdk\CLIPasswordSDK.exe>" GetPassword /p AppDescs.AppID=APP_ADOE-CAC /p Query="Safe=WW-TTS-EES-AUTOCNTR-AD;Folder=Root;Object=mmm.com_{userid}" /o Password'''

def executefunc():
    execution=AuditVerificationAnalysis().sentmailwithTable()
    return execution
class AuditVerificationAnalysis:

    def maildetails(self):
        detailspath=r"\\acprd01\E\3M_CAC\ERP_Quality_Review\auditVerification\maildetails.xlsx"
        data=pd.read_excel(detailspath,engine="openpyxl")
        leads=",".join(data[data.columns[0]].dropna().tolist())
        carbon=",".join(data[data.columns[1]].dropna().tolist())
        return {"toaddr":leads,"ccaddr":carbon}
        
    def failure(self):
        print("Files Absent,Downloading Failed")
        return "Files Absent,Downloading Failed"
    def unzip(self):
        global extract_dir
        global  ziped
        ziped = r"C:\Users\USSACPrd\Downloads"
        extract_dir = r"\\acprd01\E\3M_CAC\ERP_Quality_Review\sharepointDownloads"
        archive_format = "zip"
        files = os.listdir(ziped)
        presence = list()
        for filename in files:
            if filename.startswith("QualityCheck_Week"):
                presence.append(filename)
        if len(presence) == 0:
            AuditVerificationAnalysis.failure()
        else:
            for filename in files:
                if filename.startswith("QualityCheck_Week"):
                    fpath = os.path.join(ziped, filename)
                    shutil.unpack_archive(fpath, extract_dir, archive_format)
                    #print("Archive {filename} unpacked successfully.".format(filename=filename))
                    # os.remove(fpath)
                    # print("{filename} deleted successfully.".format(filename=filename))
        return {"folderpath":extract_dir}
    #unzip()

    '''
                   Audit Verifiaction

    '''
    def auditVerification(self):
        global Collection
        global jsoninputs
        today=str(datetime.datetime.today().date())
        jsonpath=r"\\acprd01\E\3M_CAC\ERP_Quality_Review\auditVerification\processArea.json"
        unzipping=AuditVerificationAnalysis().unzip()
        mainfilePath=unzipping.get("folderpath")#extract_dir#r"C:\Users\2040664\anuraj\ERP\sharepointDownloads"
        readjson = open(jsonpath)
        jsoninputs = json.load(readjson)
        Collection=dict()
        DetailCollection=dict()
        AllFolders=os.listdir(mainfilePath)
        #print(AllFolders)
        for eachFile in AllFolders:
            Collection[eachFile]=dict()
            DetailCollection[eachFile] = dict()
            #print(Collection)
            subFolders=os.listdir(os.path.join(mainfilePath,eachFile))
            #print(subFolders)
            for eachMacroFile in subFolders:
                eachmacroPath=os.path.join(os.path.join(mainfilePath,eachFile),eachMacroFile)
                #print(eachmacroPath)
                DetailCollection[eachFile][eachMacroFile]=dict()
                wb=openpyxl.load_workbook(filename=eachmacroPath, read_only=False, keep_vba=True,data_only=True)#openpyxl.load_workbook(eachmacroPath)
                Totalsheets=len(wb.sheetnames) #worksheets
                auditResult = 0
                for sheet in range(0,Totalsheets):
                    sheetname=wb.sheetnames[sheet]
                    if sheetname.startswith("IM"):
                        #print(sheetname)
                        ws=wb.worksheets[sheet]
                        #print(ws['E15'].value)
                        sum=ws['E15'].value
                        if sum!=None:
                            if int(ws['E15'].value)<=30:#0=done else pending

                                auditResult+=1
                        else:
                            auditResult+=1
                        DetailCollection[eachFile][eachMacroFile][sheetname]=sum
                Collection[eachFile][eachMacroFile]=auditResult
                wb.close()
        return {"Collection":Collection,"DetailCollection":DetailCollection}

        #print(Collection,len(Collection))
        #print(DetailCollection)
    #auditVerification()
    '''
                   Complete Verified week 

    '''

    def getAuditCompleted(self):
        textfilePath = r"\\acprd01\E\3M_CAC\ERP_Quality_Review\auditVerification\weeknum.txt"
        for weeknames,weeks in list(Collection.items()) :
            auditlist = []
            for process_area in list(weeks.values()):
                 auditlist.append(process_area)
                 #print(auditlist)
            if all(x == 0 for x in auditlist):
                with open(textfilePath, 'w') as f:
                    f.write(weeknames[-2:])
                #print(weeknames)
                break
    #getAuditCompleted()

    '''
                   Table Conversion

    '''

    #weeks=len(Collection)

    def CreateDashboard(self):
        execu=AuditVerificationAnalysis().auditVerification()
        Collection_=execu.get("Collection")
        
        firstlines='''<table  border="2pxsingleblack">
        <tr bgcolor="#D3D3D3">
        <td rowspan="2" style="text-align:center;" >SI.No</td>
        <td rowspan="2" style="text-align:center;" >Process Area</td>
        <td colspan="{titlecol}" style="text-align:center;" >Quality Review status as on {today}</td>
        </tr>'''.format(today=today,titlecol=str(len(Collection_)))
        weekstart='''<tr bgcolor="#ADD8E6">'''+"\n"
        weekhtml='''<td colspan="1" style="text-align:center;">{weekname}</td>'''+"\n"
        for i in range(0,len(Collection_)):
            weekstart+=weekhtml.format(weekname=list(Collection_.keys())[i].replace("QualityCheck_",""))
        weekstart+="</tr>"+"\n"
        secondwave='''<tr>'''+"\n"
        secondwavestart_pend='<td bgcolor="#FFCCCB" style="text-align:center;">{data}</td>'+"\n"
        secondwavestart = '<td bgcolor="#FFFFFF" style="text-align:center;">{data}</td>'+"\n"
        for processarea in range(0,len(jsoninputs)):
            #for weeknum in range(0, len(list(Collection_.keys()))):
            secondwave+="<tr>"+"\n"+secondwavestart.format(data=str(processarea+1))
            secondwave += secondwavestart.format(data=list(jsoninputs.keys())[processarea])
            for weeknum in range(0, len(list(Collection_.keys()))):
                data=Collection_[list(Collection_.keys())[weeknum]].get(jsoninputs[list(jsoninputs.keys())[processarea]],"-")

                if data==0: #=='pending':
                    secondwave += secondwavestart.format(data="done")
                elif data == "-":
                    secondwave += secondwavestart.format(data="-")
                else:
                    secondwave += secondwavestart_pend.format(data="pending")
            secondwave+="</tr>"+"\n"
        #print(firstlines+weekstart+secondwave+"</table>")
        tabledata=firstlines+weekstart+secondwave+"</table>"
        return tabledata
    '''
                Sent Mail
    '''
    global mailjsoninputs
    mailjsonpath=r"\\acprd01\E\3M_CAC\ERP_Quality_Review\auditVerification\mailinputs.json"
    mailreadjson = open(mailjsonpath)
    mailjsoninputs = json.load(mailreadjson)

    def sentmailwithTable(self):
        mailaspects=AuditVerificationAnalysis().maildetails()
        greeting=mailjsoninputs.get("greeting")
        bodymessage=mailjsoninputs.get("bodymessage")
        sign=mailjsoninputs.get("sign")
        email_header =mailjsoninputs.get("email_header")
        email_footer = mailjsoninputs.get("email_footer")
        to=mailaspects.get("toaddr")#mailjsoninputs.get("to")
        cc=mailaspects.get("ccaddr")#mailjsoninputs.get("cc","")
        bcc=mailjsoninputs.get("bcc","")
        From=mailjsoninputs.get("From")
        mailsubject=mailjsoninputs.get("mailsubject")
        attachments_Present=mailjsoninputs.get("attachments_Present","")
        signstart=mailjsoninputs.get("signstart")
        connectionmail=mailjsoninputs.get("connectionmail")
        smtp_server = mailjsoninputs.get("smtp_server")
        smtp_port =mailjsoninputs.get("smtp_port")
        
        attachmentpath=""
        bodydata=str()
        if attachments_Present =="true":
            attachments = attachmentpath 
        else:
            attachments=""
        html_table = AuditVerificationAnalysis().CreateDashboard()
        html_file = '''

        <!DOCTYPE html>
        <html>
        <head>
        <style>
        table {
          border-style:ridge;
          border-color:#000000;
          background-color:#000000;
          border= 1px solid;
          border-collapse: collapse;
          width: 100%;
        }

        table th {
          border : 1px solid #000000;
          padding: 6px;
          font-family: Helvetica, Arial, Helvetica;
          font-size: 12px;
        }

        table td{
          border : 1px solid #000000;
        }


        .header {
          color: white;
          background-color: green;
          border-bottom:1pt solid green;
        }

        .text_column {
          text-align: right;
        }

        .number_column {
          text-align: right;
        }

        .even_row {
          background-color: #f2f2f2;
        }

        </style>
        </head>

        <body>
          <h1></h1>
          <body style="font-family:Times New Roman">
          <br/><img src='cid:image1'<br/>
          <br>
          <br>
          <br /><font face='Times New Roman'>'''+greeting+'''</a></font><br/>

          <br /><font face='Times New Roman'>''' + bodymessage + ''' </b></font><br/>
          <br>
          <br>


        <div style="overflow-x:auto;">

        '''+html_table+'''

        </div>
        <br /><font face='Times New Roman'>'''+signstart+'''</a></font><br/>
        <br /><font face='Times New Roman'>'''+sign+''' </a></font><br/>
        <br>
        <br>
        <br/><img src='cid:image3'<br/>

        </div>
        </body>
        </html>
        '''
        msgRoot = MIMEMultipart('related')
        msgRoot['Subject'] = mailsubject.format(date=today)
        msgRoot['From'] = From
        msgRoot['Cc'] = cc
        msgRoot['To'] = to
        msgRoot['Bcc'] = bcc
        msgRoot.preamble = '====================================================='
        msgAlternative = MIMEMultipart('alternative')
        msgRoot.attach(msgAlternative)
        msgText = MIMEText('Please find ')
        msgAlternative.attach(msgText)
        msgText = MIMEText(html_file, 'html')
        msgAlternative.attach(msgText)
        msgAlternative.attach(msgText)
        fp = open(email_header, 'rb')  # "//acdev01/3M_CAC/IPM_FSM/Mail_elements/head.png"
        # fp2 = open(sys.argv[7], 'rb')#"//acdev01/3M_CAC/IPM_FSM/Mail_elements/new.png"
        fp3 = open(email_footer, 'rb')  # "//acdev01/3M_CAC/IPM_FSM/Mail_elements/footer.png"
        msgImage = MIMEImage(fp.read())
        # msgImage1 = MIMEImage(fp2.read())
        msgImage2 = MIMEImage(fp3.read())
        fp.close()
        fp3.close()
        msgImage.add_header('Content-ID', '<image1>')
        msgImage2.add_header('Content-ID', '<image3>')
        msgRoot.attach(msgImage)
        msgRoot.attach(msgImage2)
        filepaths = []
        if attachments_Present == "true":
            attachments = str(attachments).split(",")
            for f in attachments:
                with open(f, "rb") as file:
                    part = MIMEApplication(file.read(), Name=basename(f))
                    part["Content-Disposition"] = 'attachment;filename="%s"' % basename(f)
                    msgRoot.attach(part)
        smtp = smtplib.SMTP()
        smtp.connect(connectionmail)
        # smtp.sendmail(From,To, msgRoot.as_string())
        smtp.send_message(msgRoot)
        smtp.quit()
        deletion=AuditVerificationAnalysis().deleteFiles()
        print("Email is sent successfully")
        return "Email is sent successfully"+ deletion
    #sentmailwithTable()
    '''
                  Delete Files
    '''
    def deleteFiles(self):
        for file in os.listdir(extract_dir):
            shutil.rmtree(os.path.join(extract_dir,file))
        for file in os.listdir(ziped):
            if "QualityCheck_Week" in file:
                os.remove(os.path.join(ziped,file))
        return "Files removed successfully"
        
#executefunc()





            



c='''{'QualityCheck_Week_35': {'US_IT-SAPOne-Application.xlsm': 1, 'US_NGG-APPL-Support.xlsm': 0, 'WW_HTR-LSO-Support.xlsm': 0, 'WW_L2-GOS-FTS.xlsm': 5, 'WW_L2-
GOS-ICO.xlsm': 5, 'WW_L2-GOS-IM-INTG.xlsm': 2, 'WW_L2-GOS-LEX-GTS.xlsm': 0, 'WW_L2-GOS-LEX-TM.xlsm': 0, 'WW_L2-GOS-LEX-WHSE.xlsm': 6, 'WW_L2-GOS-MTD.xlsm':
 0, 'WW_L2-GOS-MTO.xlsm': 6, 'WW_L2-GOS-OTC.xlsm': 0, 'WW_L2-GOS-RTR.xlsm': 0, 'WW_L2-GOS-STP.xlsm': 0, 'WW_L2-RunSAP-SO-EHS.xlsm': 0}, 'QualityCheck_Week_
36': {'US_IT-SAPOne-Application.xlsm': 0, 'US_NGG-APPL-Support.xlsm': 0, 'WW_HTR-LSO-Support.xlsm': 0, 'WW_L2-GOS-FTS.xlsm': 0, 'WW_L2-GOS-ICO.xlsm': 6, 'W
W_L2-GOS-IM-INTG.xlsm': 4, 'WW_L2-GOS-LEX-GTS.xlsm': 0, 'WW_L2-GOS-LEX-TM.xlsm': 0, 'WW_L2-GOS-LEX-WHSE.xlsm': 6, 'WW_L2-GOS-MTD.xlsm': 5, 'WW_L2-GOS-MTO.x
lsm': 0, 'WW_L2-GOS-OTC.xlsm': 0, 'WW_L2-GOS-RTR.xlsm': 0, 'WW_L2-GOS-STP.xlsm': 0, 'WW_L2-RunSAP-SO-EHS.xlsm': 0, 'WW_L2-RunSAPIT-WRICEF.xlsm': 0}'''

