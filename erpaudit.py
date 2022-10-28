'''

Author      :      Anuraj Pilanku

Use Case    :     ERP Audit Verification Automation

'''
import openpyxl
import os
import json
import datetime
import shutil
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from os.path import basename
import pandas as pd
import sys
from datetime import date
import datetime
import time
from exchangelib.protocol import BaseProtocol
from exchangelib import Credentials, Account, DELEGATE, Configuration, FileAttachment
import requests
from bs4 import BeautifulSoup, BeautifulStoneSoup

today=str(datetime.datetime.today().date())
passwordUpdate='''"C:\Program Files (x86)\CyberArk\ApplicationPasswordSdk\CLIPasswordSDK.exe>" GetPassword /p AppDescs.AppID=APP_ADOE-CAC /p Query="Safe=WW-TTS-EES-AUTOCNTR-AD;Folder=Root;Object=mmm.com_{userid}" /o Password'''

'''
               File Unzip

'''
class ProxyAdapter(requests.adapters.HTTPAdapter):
    def send(self, *args, **kwargs):
        kwargs["proxies"] = ProxyAdapter.proxies
        return super(ProxyAdapter, self).send(*args, **kwargs)


class MonitorMail:
    def _connect(self):  # (self, aeuser, aepassword, aeci, aeparameters):
        '''
        Method for creating connection
        and configuration to monitor mailbox
        '''
        o365mailid = "USSACDev@mmm.com"
        mailusername = "USSACDev"#jsoninputs["user1"].strip()
        proxy = {}
        proxy.update({})  # (aeci.get("proxy", {}))
        ProxyAdapter.proxies = proxy
        BaseProtocol.HTTP_ADAPTER_CLS = ProxyAdapter
        smtp_address = "USSACDev@mmm.com"  # "mailserv.mmm.com"#aeci.get("primarysmtpaddress", aeuser)
        if not smtp_address:
            smtp_address = "USSACDev@mmm.com"  # aeuser
        server = "Outlook.Office365.com"  # "mailserv.mmm.com"# aeci.get("server", "Outlook.Office365.com")
        if not server:
            server = "Outlook.Office365.com"
        credentials = Credentials(username=o365mailid, password=os.popen(passwordUpdate.format(
            userid=mailusername)).read().strip())  # ",=I}-)7(}4Uf_%</Mrl/")  # password="vLcRU0%@ZnUd+Xy(.a5w")
        config = Configuration(
            server=server,
            credentials=credentials)
        account = Account(
            primary_smtp_address=smtp_address,
            credentials=credentials,
            autodiscover=False,
            config=config,
            access_type=DELEGATE)
        return account

    def connectFolder(self):
        mailfolder = MonitorMail()._connect().root
        ourfolder = MonitorMail()._connect()

        folder_name = "ERPAuditPending"  # "ApplensSMO"#aeparameters.get("folder_name")
        if folder_name:
            my_folder = ourfolder.root / "Top of Information Store"
            for folder in folder_name.split("/"):
                my_folder = my_folder / folder
            filter_criteria = 'isRead:False'  # AND ({0})'.format(createactivitykeywords)
            items = my_folder.filter(filter_criteria)
        else:
            filter_criteria = 'isRead:False'  # AND ({0})'.format(createactivitykeywords)
            items = ourfolder.inbox.filter(filter_criteria)
        result = {}
        result["items"] = items

        filter_criteria = 'isRead:False'
        payload = {}
        mailCollections = {}
        count = 0
        for item in items.__iter__():
            soup = BeautifulSoup(item.text_body, 'html.parser')
            body = soup.text.replace('\n', "").replace('\r', "").strip()
            subject = item.subject.replace('\n', "").replace('\r', "").strip()  # ID:IM43545678 CODE:CACHPSMO365AUTO
            mailaddress = item.author.email_address.replace('\n', "").replace('\r', "").strip()
            count += 1
            mailCollections["mail_{0}".format(str(count))] = {"subject": subject.strip(), "mailaddress": mailaddress.strip(),
                                                              "body": body.strip()}
            item.is_read = True
            item.save()
            if len(mailCollections) != 0:                                                     
                unzip()
                auditVerification()
                getAuditCompleted()
                sentmailwithTable()
            # to make mail as read!!
        
class erpaudit:
    def run(self):
        runfile=MonitorMail()
        return runfile.connectFolder()

def failure():
    print("Files Absent,Downloading Failed")
    return "Files Absent,Downloading Failed"
def unzip():
    global extract_dir
    global  ziped
    ziped = r"\\acdev01\3M_CAC\ERP_Quality_Review\auditVerification\zip"
    extract_dir = r"\\acdev01\3M_CAC\ERP_Quality_Review\auditVerification\unzip"
    archive_format = "zip"
    files = os.listdir(ziped)
    presence = list()
    for filename in files:
        if filename.startswith("QualityCheck_Week"):
            presence.append(filename)
    if len(presence) == 0:
        failure()
    else:
        for filename in files:
            if filename.startswith("QualityCheck_Week"):
                fpath = os.path.join(ziped, filename)
                shutil.unpack_archive(fpath, extract_dir, archive_format)
                #print("Archive {filename} unpacked successfully.".format(filename=filename))
                # os.remove(fpath)
                # print("{filename} deleted successfully.".format(filename=filename))
#unzip()

'''
               Audit Verifiaction

'''
def auditVerification():
    global Collection
    global jsoninputs
    today=str(datetime.datetime.today().date())
    jsonpath=r"\\acdev01\3M_CAC\ERP_Quality_Review\auditVerification\processArea.json"
    mainfilePath=extract_dir#r"C:\Users\2040664\anuraj\ERP\sharepointDownloads"
    readjson = open(jsonpath)
    jsoninputs = json.load(readjson)
    Collection=dict()
    DetailCollection=dict()
    AllFolders=os.listdir(mainfilePath)
    #print(AllFolders)
    for eachFile in AllFolders:
        Collection[eachFile]=dict()
        DetailCollection[eachFile] = dict()
        #print(Collection)
        subFolders=os.listdir(os.path.join(mainfilePath,eachFile))
        #print(subFolders)
        for eachMacroFile in subFolders:
            eachmacroPath=os.path.join(os.path.join(mainfilePath,eachFile),eachMacroFile)
            #print(eachmacroPath)
            DetailCollection[eachFile][eachMacroFile]=dict()
            wb=openpyxl.load_workbook(filename=eachmacroPath, read_only=False, keep_vba=True,data_only=True)#openpyxl.load_workbook(eachmacroPath)
            Totalsheets=len(wb.sheetnames) #worksheets
            auditResult = 0
            for sheet in range(0,Totalsheets):
                sheetname=wb.sheetnames[sheet]
                if sheetname.startswith("IM"):
                    #print(sheetname)
                    ws=wb.worksheets[sheet]
                    #print(ws['E15'].value)
                    sum=ws['E15'].value
                    if sum!=None:
                        if int(ws['E15'].value)<=30:#0=done else pending

                            auditResult+=1
                    else:
                        auditResult+=1
                    DetailCollection[eachFile][eachMacroFile][sheetname]=sum
            Collection[eachFile][eachMacroFile]=auditResult
            wb.close()

    #print(Collection,len(Collection))
    #print(DetailCollection)
#auditVerification()
'''
               Complete Verified week 

'''
def getAuditCompleted():
    textfilePath = r"\\acdev01\3M_CAC\ERP_Quality_Review\auditVerification\weeknum.txt"
    for weeknames,weeks in list(Collection.items()) :
        auditlist = []
        for process_area in list(weeks.values()):
             auditlist.append(process_area)
             #print(auditlist)
        if all(x == 0 for x in auditlist):
            with open(textfilePath, 'w') as f:
                f.write(weeknames[-2:])
            #print(weeknames)
            break
#getAuditCompleted()

'''
               Table Conversion

'''

#weeks=len(Collection)

def CreateDashboard():
    firstlines='''<table  border="2pxsingleblack">
    <tr bgcolor="#D3D3D3">
    <td rowspan="2" style="text-align:center;" >SI.No</td>
    <td rowspan="2" style="text-align:center;" >Process Area</td>
    <td colspan="3" style="text-align:center;" >Quality Review status as on {today}</td>
    </tr>'''.format(today=today)
    weekstart='''<tr bgcolor="#ADD8E6">'''+"\n"
    weekhtml='''<td colspan="1" style="text-align:center;">{weekname}</td>'''+"\n"
    for i in range(0,len(Collection)):
        weekstart+=weekhtml.format(weekname=list(Collection.keys())[i].replace("QualityCheck_",""))
    weekstart+="</tr>"+"\n"
    secondwave='''<tr>'''+"\n"
    secondwavestart_pend='<td bgcolor="#FFCCCB" style="text-align:center;">{data}</td>'+"\n"
    secondwavestart = '<td bgcolor="#FFFFFF" style="text-align:center;">{data}</td>'+"\n"
    for processarea in range(0,len(jsoninputs)):
        #for weeknum in range(0, len(list(Collection.keys()))):
        secondwave+="<tr>"+"\n"+secondwavestart.format(data=str(processarea+1))
        secondwave += secondwavestart.format(data=list(jsoninputs.keys())[processarea])
        for weeknum in range(0, len(list(Collection.keys()))):
            data=Collection[list(Collection.keys())[weeknum]].get(jsoninputs[list(jsoninputs.keys())[processarea]],"-")

            if data==0: #=='pending':
                secondwave += secondwavestart.format(data="done")
            elif data == "-":
                secondwave += secondwavestart.format(data="-")
            else:
                secondwave += secondwavestart_pend.format(data="pending")
        secondwave+="</tr>"+"\n"
    #print(firstlines+weekstart+secondwave+"</table>")
    return firstlines+weekstart+secondwave+"</table>"
'''
            Sent Mail
'''
mailjsonpath=r"\\acdev01\3M_CAC\ERP_Quality_Review\auditVerification\mailinputs.json"
mailreadjson = open(mailjsonpath)
mailjsoninputs = json.load(mailreadjson)
greeting=mailjsoninputs.get("greeting")
bodymessage=mailjsoninputs.get("bodymessage")
sign=mailjsoninputs.get("sign")
email_header =mailjsoninputs.get("email_header")
email_footer = mailjsoninputs.get("email_footer")
to=mailjsoninputs.get("to")
cc=mailjsoninputs.get("cc","")
bcc=mailjsoninputs.get("bcc","")
From=mailjsoninputs.get("From")
mailsubject=mailjsoninputs.get("mailsubject")
attachments_Present=mailjsoninputs.get("attachments_Present","")
signstart=mailjsoninputs.get("signstart")
connectionmail=mailjsoninputs.get("connectionmail")
smtp_server = mailjsoninputs.get("smtp_server")
smtp_port =mailjsoninputs.get("smtp_port")

def sentmailwithTable():
    attachmentpath=""
    bodydata=str()
    if attachments_Present =="true":
        attachments = attachmentpath # r"C:\Users\USSACDev\Desktop\anuraj\nazia\nazia.xlsx"
    else:
        attachments=""
    html_table = CreateDashboard()
    html_file = '''

    <!DOCTYPE html>
    <html>
    <head>
    <style>
    table {
      border-style:ridge;
      border-color:#000000;
      background-color:#000000;
      border= 1px solid;
      border-collapse: collapse;
      width: 100%;
    }

    table th {
      border : 1px solid #000000;
      padding: 6px;
      font-family: Helvetica, Arial, Helvetica;
      font-size: 12px;
    }

    table td{
      border : 1px solid #000000;
    }


    .header {
      color: white;
      background-color: green;
      border-bottom:1pt solid green;
    }

    .text_column {
      text-align: right;
    }

    .number_column {
      text-align: right;
    }

    .even_row {
      background-color: #f2f2f2;
    }

    </style>
    </head>

    <body>
      <h1></h1>
      <body style="font-family:Times New Roman">
      <br/><img src='cid:image1'<br/>
      <br>
      <br>
      <br /><font face='Times New Roman'>'''+greeting+'''</a></font><br/>

      <br /><font face='Times New Roman'>''' + bodymessage + ''' </b></font><br/>
      <br>
      <br>


    <div style="overflow-x:auto;">

    '''+html_table+'''

    </div>
    <br /><font face='Times New Roman'>'''+signstart+'''</a></font><br/>
    <br /><font face='Times New Roman'>'''+sign+''' </a></font><br/>
    <br>
    <br>
    <br/><img src='cid:image3'<br/>

    </div>
    </body>
    </html>
    '''
    msgRoot = MIMEMultipart('related')
    msgRoot['Subject'] = mailsubject.format(date=today)
    msgRoot['From'] = From
    msgRoot['Cc'] = cc
    msgRoot['To'] = to
    msgRoot['Bcc'] = bcc
    msgRoot.preamble = '====================================================='
    msgAlternative = MIMEMultipart('alternative')
    msgRoot.attach(msgAlternative)
    msgText = MIMEText('Please find ')
    msgAlternative.attach(msgText)
    msgText = MIMEText(html_file, 'html')
    msgAlternative.attach(msgText)
    msgAlternative.attach(msgText)
    fp = open(email_header, 'rb')  # "//acdev01/3M_CAC/IPM_FSM/Mail_elements/head.png"
    # fp2 = open(sys.argv[7], 'rb')#"//acdev01/3M_CAC/IPM_FSM/Mail_elements/new.png"
    fp3 = open(email_footer, 'rb')  # "//acdev01/3M_CAC/IPM_FSM/Mail_elements/footer.png"
    msgImage = MIMEImage(fp.read())
    # msgImage1 = MIMEImage(fp2.read())
    msgImage2 = MIMEImage(fp3.read())
    fp.close()
    fp3.close()
    msgImage.add_header('Content-ID', '<image1>')
    msgImage2.add_header('Content-ID', '<image3>')
    msgRoot.attach(msgImage)
    msgRoot.attach(msgImage2)
    filepaths = []
    if attachments_Present == "true":
        attachments = str(attachments).split(",")
        for f in attachments:
            with open(f, "rb") as file:
                part = MIMEApplication(file.read(), Name=basename(f))
                part["Content-Disposition"] = 'attachment;filename="%s"' % basename(f)
                msgRoot.attach(part)
    smtp = smtplib.SMTP()
    smtp.connect(connectionmail)
    # smtp.sendmail(From,To, msgRoot.as_string())
    smtp.send_message(msgRoot)
    smtp.quit()
    print("Email is sent successfully")
#sentmailwithTable()
'''
              Delete Files
'''
def deleteFiles():
    for file in os.listdir(extract_dir):
        os.remove(os.path.join(extract_dir,file))
    for file in os.listdir(ziped):
        os.remove(os.path.join(ziped,file))

'''
              Monitor Mail
'''



            



c='''{'QualityCheck_Week_35': {'US_IT-SAPOne-Application.xlsm': 1, 'US_NGG-APPL-Support.xlsm': 0, 'WW_HTR-LSO-Support.xlsm': 0, 'WW_L2-GOS-FTS.xlsm': 5, 'WW_L2-
GOS-ICO.xlsm': 5, 'WW_L2-GOS-IM-INTG.xlsm': 2, 'WW_L2-GOS-LEX-GTS.xlsm': 0, 'WW_L2-GOS-LEX-TM.xlsm': 0, 'WW_L2-GOS-LEX-WHSE.xlsm': 6, 'WW_L2-GOS-MTD.xlsm':
 0, 'WW_L2-GOS-MTO.xlsm': 6, 'WW_L2-GOS-OTC.xlsm': 0, 'WW_L2-GOS-RTR.xlsm': 0, 'WW_L2-GOS-STP.xlsm': 0, 'WW_L2-RunSAP-SO-EHS.xlsm': 0}, 'QualityCheck_Week_
36': {'US_IT-SAPOne-Application.xlsm': 0, 'US_NGG-APPL-Support.xlsm': 0, 'WW_HTR-LSO-Support.xlsm': 0, 'WW_L2-GOS-FTS.xlsm': 0, 'WW_L2-GOS-ICO.xlsm': 6, 'W
W_L2-GOS-IM-INTG.xlsm': 4, 'WW_L2-GOS-LEX-GTS.xlsm': 0, 'WW_L2-GOS-LEX-TM.xlsm': 0, 'WW_L2-GOS-LEX-WHSE.xlsm': 6, 'WW_L2-GOS-MTD.xlsm': 5, 'WW_L2-GOS-MTO.x
lsm': 0, 'WW_L2-GOS-OTC.xlsm': 0, 'WW_L2-GOS-RTR.xlsm': 0, 'WW_L2-GOS-STP.xlsm': 0, 'WW_L2-RunSAP-SO-EHS.xlsm': 0, 'WW_L2-RunSAPIT-WRICEF.xlsm': 0}'''

